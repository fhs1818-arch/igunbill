#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
굿옥션 경매 관심물건 × 국토부 아파트 매매 실거래가 조회
- 굿옥션에서 내려받은 HTML-XLS(euc-kr) 파일을 읽고
- 각 물건 주소에서 법정동코드(5자리)를 자동 추출하여
- 국토부 아파트 매매 실거래가 API로 최근 6개월 조회 후
- 결과를 물건정보 + 실거래가 나란히 엑셀로 저장

필요 패키지:  pip install pandas requests openpyxl lxml
"""

import io
import os
import re
import sys
import time
import xml.etree.ElementTree as ET
from datetime import datetime

import pandas as pd
import requests
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# 설정
# ─────────────────────────────────────────────────────────────────────────────
SERVICE_KEY = "81e2448545772d4b5f9ad3039b23b81af42ecaeee706ae4c6fb6d1241919e08c"
API_URL = "http://apis.data.go.kr/1613000/RTMSDataSvcAptTrade/getRTMSDataSvcAptTrade"
RECENT_MONTHS = 6       # 조회 월 수
MAX_TRADES_PER_ITEM = 30  # 물건당 최대 실거래 표시 건수
API_DELAY = 0.15        # 요청 간격(초) - 과부하 방지

# ─────────────────────────────────────────────────────────────────────────────
# 시도 약칭 → 정식명
# ─────────────────────────────────────────────────────────────────────────────
SIDO_ABBR = {
    '서울': '서울', '부산': '부산', '대구': '대구', '인천': '인천',
    '광주': '광주', '대전': '대전', '울산': '울산', '세종': '세종',
    '경기': '경기', '강원': '강원', '충북': '충북', '충남': '충남',
    '전북': '전북', '전남': '전남', '경북': '경북', '경남': '경남',
    '제주': '제주',
    # 정식명도 허용
    '서울특별시': '서울', '부산광역시': '부산', '대구광역시': '대구',
    '인천광역시': '인천', '광주광역시': '광주', '대전광역시': '대전',
    '울산광역시': '울산', '세종특별자치시': '세종', '경기도': '경기',
    '강원도': '강원', '강원특별자치도': '강원',
    '충청북도': '충북', '충청남도': '충남',
    '전라북도': '전북', '전북특별자치도': '전북',
    '전라남도': '전남', '경상북도': '경북', '경상남도': '경남',
    '제주특별자치도': '제주',
}

# ─────────────────────────────────────────────────────────────────────────────
# 시군구 → 법정동코드 5자리  (키: "시도약칭 시군구명")
# ─────────────────────────────────────────────────────────────────────────────
LAWD_CD_MAP = {
    # ── 서울특별시 ──
    '서울 종로구': '11110', '서울 중구': '11140', '서울 용산구': '11170',
    '서울 성동구': '11200', '서울 광진구': '11215', '서울 동대문구': '11230',
    '서울 중랑구': '11260', '서울 성북구': '11290', '서울 강북구': '11305',
    '서울 도봉구': '11320', '서울 노원구': '11350', '서울 은평구': '11380',
    '서울 서대문구': '11410', '서울 마포구': '11440', '서울 양천구': '11470',
    '서울 강서구': '11500', '서울 구로구': '11530', '서울 금천구': '11545',
    '서울 영등포구': '11560', '서울 동작구': '11590', '서울 관악구': '11620',
    '서울 서초구': '11650', '서울 강남구': '11680', '서울 송파구': '11710',
    '서울 강동구': '11740',
    # ── 부산광역시 ──
    '부산 중구': '26110', '부산 서구': '26140', '부산 동구': '26170',
    '부산 영도구': '26200', '부산 부산진구': '26230', '부산 동래구': '26260',
    '부산 남구': '26290', '부산 북구': '26320', '부산 해운대구': '26350',
    '부산 사하구': '26380', '부산 금정구': '26410', '부산 강서구': '26440',
    '부산 연제구': '26470', '부산 수영구': '26500', '부산 사상구': '26530',
    '부산 기장군': '26710',
    # ── 대구광역시 ──
    '대구 중구': '27110', '대구 동구': '27140', '대구 서구': '27170',
    '대구 남구': '27200', '대구 북구': '27230', '대구 수성구': '27260',
    '대구 달서구': '27290', '대구 달성군': '27710', '대구 군위군': '27720',
    # ── 인천광역시 ──
    '인천 중구': '28110', '인천 동구': '28140', '인천 미추홀구': '28177',
    '인천 연수구': '28185', '인천 남동구': '28200', '인천 부평구': '28237',
    '인천 계양구': '28245', '인천 서구': '28260',
    '인천 강화군': '28710', '인천 옹진군': '28720',
    # ── 광주광역시 ──
    '광주 동구': '29110', '광주 서구': '29140', '광주 남구': '29155',
    '광주 북구': '29170', '광주 광산구': '29200',
    # ── 대전광역시 ──
    '대전 동구': '30110', '대전 중구': '30140', '대전 서구': '30170',
    '대전 유성구': '30200', '대전 대덕구': '30230',
    # ── 울산광역시 ──
    '울산 중구': '31110', '울산 남구': '31140', '울산 동구': '31170',
    '울산 북구': '31200', '울산 울주군': '31710',
    # ── 세종특별자치시 ──
    '세종 세종시': '36110',
    # ── 경기도 ──
    '경기 수원시 장안구': '41111', '경기 수원시 권선구': '41113',
    '경기 수원시 팔달구': '41115', '경기 수원시 영통구': '41117',
    '경기 성남시 수정구': '41131', '경기 성남시 중원구': '41133',
    '경기 성남시 분당구': '41135',
    '경기 의정부시': '41150',
    '경기 안양시 만안구': '41171', '경기 안양시 동안구': '41173',
    '경기 부천시': '41190', '경기 광명시': '41210', '경기 평택시': '41220',
    '경기 동두천시': '41250',
    '경기 안산시 상록구': '41271', '경기 안산시 단원구': '41273',
    '경기 고양시 덕양구': '41281', '경기 고양시 일산동구': '41285',
    '경기 고양시 일산서구': '41287',
    '경기 과천시': '41290', '경기 구리시': '41310', '경기 남양주시': '41360',
    '경기 오산시': '41370', '경기 시흥시': '41390', '경기 군포시': '41410',
    '경기 의왕시': '41430', '경기 하남시': '41450',
    '경기 용인시 처인구': '41461', '경기 용인시 기흥구': '41463',
    '경기 용인시 수지구': '41465',
    '경기 파주시': '41480', '경기 이천시': '41500', '경기 안성시': '41550',
    '경기 김포시': '41570', '경기 화성시': '41590', '경기 광주시': '41610',
    '경기 양주시': '41630', '경기 포천시': '41650', '경기 여주시': '41670',
    '경기 연천군': '41800', '경기 가평군': '41820', '경기 양평군': '41830',
    # ── 강원도 / 강원특별자치도 ──
    '강원 춘천시': '42110', '강원 원주시': '42130', '강원 강릉시': '42150',
    '강원 동해시': '42170', '강원 태백시': '42190', '강원 속초시': '42210',
    '강원 삼척시': '42230', '강원 홍천군': '42720', '강원 횡성군': '42730',
    '강원 영월군': '42750', '강원 평창군': '42760', '강원 정선군': '42770',
    '강원 철원군': '42780', '강원 화천군': '42790', '강원 양구군': '42800',
    '강원 인제군': '42810', '강원 고성군': '42820', '강원 양양군': '42830',
    # ── 충청북도 ──
    '충북 청주시 상당구': '43111', '충북 청주시 서원구': '43112',
    '충북 청주시 흥덕구': '43113', '충북 청주시 청원구': '43114',
    '충북 충주시': '43130', '충북 제천시': '43150',
    '충북 보은군': '43720', '충북 옥천군': '43730', '충북 영동군': '43740',
    '충북 증평군': '43745', '충북 진천군': '43750', '충북 괴산군': '43760',
    '충북 음성군': '43770', '충북 단양군': '43800',
    # ── 충청남도 ──
    '충남 천안시 동남구': '44131', '충남 천안시 서북구': '44133',
    '충남 공주시': '44150', '충남 보령시': '44180', '충남 아산시': '44200',
    '충남 서산시': '44210', '충남 논산시': '44230', '충남 계룡시': '44250',
    '충남 당진시': '44270',
    '충남 금산군': '44710', '충남 부여군': '44760', '충남 서천군': '44770',
    '충남 청양군': '44790', '충남 홍성군': '44800', '충남 예산군': '44810',
    '충남 태안군': '44825',
    # ── 전라북도 / 전북특별자치도 ──
    '전북 전주시 완산구': '45111', '전북 전주시 덕진구': '45113',
    '전북 군산시': '45130', '전북 익산시': '45140', '전북 정읍시': '45180',
    '전북 남원시': '45190', '전북 김제시': '45210',
    '전북 완주군': '45710', '전북 진안군': '45720', '전북 무주군': '45730',
    '전북 장수군': '45740', '전북 임실군': '45750', '전북 순창군': '45770',
    '전북 고창군': '45790', '전북 부안군': '45800',
    # ── 전라남도 ──
    '전남 목포시': '46110', '전남 여수시': '46130', '전남 순천시': '46150',
    '전남 나주시': '46170', '전남 광양시': '46230',
    '전남 담양군': '46710', '전남 곡성군': '46720', '전남 구례군': '46730',
    '전남 고흥군': '46770', '전남 보성군': '46780', '전남 화순군': '46790',
    '전남 장흥군': '46800', '전남 강진군': '46810', '전남 해남군': '46820',
    '전남 영암군': '46830', '전남 무안군': '46840', '전남 함평군': '46860',
    '전남 영광군': '46870', '전남 장성군': '46880', '전남 완도군': '46890',
    '전남 진도군': '46900', '전남 신안군': '46910',
    # ── 경상북도 ──
    '경북 포항시 남구': '47111', '경북 포항시 북구': '47113',
    '경북 경주시': '47130', '경북 김천시': '47150', '경북 안동시': '47170',
    '경북 구미시': '47190', '경북 영주시': '47210', '경북 영천시': '47230',
    '경북 상주시': '47250', '경북 문경시': '47280', '경북 경산시': '47290',
    '경북 의성군': '47730', '경북 청송군': '47750', '경북 영양군': '47760',
    '경북 영덕군': '47770', '경북 청도군': '47820', '경북 고령군': '47830',
    '경북 성주군': '47840', '경북 칠곡군': '47850', '경북 예천군': '47900',
    '경북 봉화군': '47920', '경북 울진군': '47930', '경북 울릉군': '47940',
    # ── 경상남도 ──
    '경남 창원시 의창구': '48121', '경남 창원시 성산구': '48123',
    '경남 창원시 마산합포구': '48125', '경남 창원시 마산회원구': '48127',
    '경남 창원시 진해구': '48129',
    '경남 진주시': '48170', '경남 통영시': '48220', '경남 사천시': '48240',
    '경남 김해시': '48250', '경남 밀양시': '48270', '경남 거제시': '48310',
    '경남 양산시': '48330',
    '경남 의령군': '48720', '경남 함안군': '48730', '경남 창녕군': '48740',
    '경남 고성군': '48820', '경남 남해군': '48840', '경남 하동군': '48850',
    '경남 산청군': '48860', '경남 함양군': '48870', '경남 거창군': '48880',
    '경남 합천군': '48890',
    # ── 제주특별자치도 ──
    '제주 제주시': '50110', '제주 서귀포시': '50130',
}


# ─────────────────────────────────────────────────────────────────────────────
# 주소 → 법정동코드 추출
# ─────────────────────────────────────────────────────────────────────────────
def get_lawd_cd(address: str):
    """주소 문자열에서 법정동코드 5자리와 시군구명을 반환."""
    addr = re.sub(r'\s+', ' ', address.strip())

    # 시도 약칭 찾기
    sido = None
    for key in sorted(SIDO_ABBR, key=len, reverse=True):
        if key in addr:
            sido = SIDO_ABBR[key]
            break
    if not sido:
        return None, None

    # 긴 키부터 매칭 (예: "용인시 기흥구" 가 "용인시"보다 먼저)
    for map_key in sorted(LAWD_CD_MAP, key=len, reverse=True):
        prefix, sigungu = map_key.split(' ', 1)
        if prefix != sido:
            continue
        if sigungu in addr:
            return LAWD_CD_MAP[map_key], sigungu

    return None, None


# ─────────────────────────────────────────────────────────────────────────────
# 날짜 유틸
# ─────────────────────────────────────────────────────────────────────────────
def recent_months(n: int = 6) -> list[str]:
    """현재 시점 기준 최근 n개월 목록 (YYYYMM 형식)."""
    result = []
    now = datetime.now()
    y, m = now.year, now.month
    for _ in range(n):
        result.append(f"{y}{m:02d}")
        m -= 1
        if m == 0:
            m, y = 12, y - 1
    return result


# ─────────────────────────────────────────────────────────────────────────────
# 국토부 API
# ─────────────────────────────────────────────────────────────────────────────
def query_apt_trade_month(lawd_cd: str, deal_ymd: str) -> list[dict]:
    """국토부 아파트 매매 실거래가 API – 특정 월 조회."""
    params = {
        'serviceKey': SERVICE_KEY,
        'LAWD_CD': lawd_cd,
        'DEAL_YMD': deal_ymd,
        'numOfRows': 1000,
        'pageNo': 1,
    }
    try:
        resp = requests.get(API_URL, params=params, timeout=20)
        resp.raise_for_status()

        root = ET.fromstring(resp.content)
        code = root.findtext('.//resultCode', '')
        if code not in ('', '00', '0000', 'INFO-000'):
            msg = root.findtext('.//resultMsg', '알 수 없는 오류')
            print(f"    ⚠ API [{deal_ymd}] {code}: {msg}")
            return []

        return [
            {child.tag.strip(): (child.text or '').strip() for child in item}
            for item in root.findall('.//item')
        ]

    except ET.ParseError:
        print(f"    ⚠ XML 파싱 실패 ({deal_ymd})")
        return []
    except Exception as e:
        print(f"    ⚠ 요청 실패 ({deal_ymd}): {e}")
        return []


def fetch_recent_trades(lawd_cd: str) -> list[dict]:
    """최근 RECENT_MONTHS 개월치 실거래가 수집."""
    trades = []
    for ym in recent_months(RECENT_MONTHS):
        batch = query_apt_trade_month(lawd_cd, ym)
        trades.extend(batch)
        time.sleep(API_DELAY)
    return trades


# ─────────────────────────────────────────────────────────────────────────────
# 필터 & 포맷
# ─────────────────────────────────────────────────────────────────────────────
def extract_dong(address: str) -> str:
    """주소에서 읍면동 이름 추출 (예: '개포동' → '개포')."""
    m = re.search(r'(\S+?)[동읍면리]\b', address)
    return m.group(1) if m else ''


def filter_by_dong(trades: list[dict], address: str) -> list[dict]:
    """주소의 동 이름 앞 2글자로 실거래 필터링."""
    dong_prefix = extract_dong(address)
    if len(dong_prefix) < 2:
        return trades
    return [t for t in trades if dong_prefix in t.get('umdNm', '')]


def sort_trades(trades: list[dict]) -> list[dict]:
    """거래일 내림차순 정렬."""
    return sorted(
        trades,
        key=lambda t: (
            t.get('dealYear', ''),
            t.get('dealMonth', '').zfill(2),
            t.get('dealDay', '').zfill(2),
        ),
        reverse=True,
    )


def fmt_trade(t: dict) -> dict:
    """API 응답 항목을 출력용 딕셔너리로 변환."""
    y = t.get('dealYear', '')
    mo = t.get('dealMonth', '').zfill(2)
    d = t.get('dealDay', '').zfill(2)
    return {
        '실거래_거래일': f"{y}-{mo}-{d}" if y else '',
        '실거래_아파트명': t.get('aptNm', ''),
        '실거래_읍면동': t.get('umdNm', ''),
        '실거래_지번': t.get('jibun', ''),
        '실거래_전용면적㎡': t.get('excluUseAr', ''),
        '실거래_층': t.get('floor', ''),
        '실거래_건축년도': t.get('buildYear', ''),
        '실거래_거래금액만원': t.get('dealAmount', '').replace(',', ''),
        '실거래_매수자구분': t.get('buyerGbn', ''),
    }


# ─────────────────────────────────────────────────────────────────────────────
# 굿옥션 XLS 읽기
# ─────────────────────────────────────────────────────────────────────────────
def read_goodauction_xls(path: str) -> pd.DataFrame:
    """굿옥션 HTML-XLS(euc-kr) 파일을 DataFrame으로 반환."""
    with open(path, 'rb') as f:
        raw = f.read()

    for enc in ('euc-kr', 'cp949', 'utf-8-sig', 'utf-8'):
        try:
            html = raw.decode(enc)
            break
        except (UnicodeDecodeError, LookupError):
            continue
    else:
        html = raw.decode('cp949', errors='replace')

    tables = pd.read_html(io.StringIO(html), flavor='lxml')
    if not tables:
        raise ValueError("테이블을 찾을 수 없습니다.")

    # 행이 가장 많은 테이블 선택
    df = max(tables, key=len)
    df = df.dropna(how='all').reset_index(drop=True)
    df.columns = [str(c).strip() for c in df.columns]

    # 헤더가 데이터 행에 반복되는 경우 제거
    if len(df) > 1:
        header_vals = set(df.columns)
        df = df[~df.iloc[:, 0].astype(str).isin(header_vals)].reset_index(drop=True)

    return df


def find_address_col(df: pd.DataFrame) -> str | None:
    """소재지/주소 컬럼을 자동 탐지."""
    keywords = ['소재지', '주소', '물건소재지', '위치']
    for col in df.columns:
        if any(kw in str(col) for kw in keywords):
            return col
    return None


# ─────────────────────────────────────────────────────────────────────────────
# 엑셀 저장
# ─────────────────────────────────────────────────────────────────────────────
def save_excel(result_df: pd.DataFrame, out_path: str, auction_col_count: int):
    """결과를 서식 적용한 엑셀로 저장."""
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        result_df.to_excel(writer, index=False, sheet_name='실거래가분석')
        ws = writer.sheets['실거래가분석']

        blue_fill  = PatternFill(fill_type='solid', fgColor='4472C4')
        orange_fill = PatternFill(fill_type='solid', fgColor='ED7D31')
        white_font = Font(bold=True, color='FFFFFF')

        for ci, cell in enumerate(ws[1], 1):
            cell.font = white_font
            cell.fill = blue_fill if ci <= auction_col_count else orange_fill

        # 열 너비 자동 조정
        for ci, col_cells in enumerate(ws.columns, 1):
            width = max(
                (len(str(c.value)) for c in col_cells if c.value is not None),
                default=8,
            )
            ws.column_dimensions[get_column_letter(ci)].width = min(width + 2, 40)

        ws.freeze_panes = 'A2'


# ─────────────────────────────────────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────────────────────────────────────
def main():
    print("=" * 58)
    print("  굿옥션 경매물건 × 국토부 아파트 매매 실거래가 조회")
    print("=" * 58)

    if len(sys.argv) > 1:
        file_path = sys.argv[1].strip().strip('"').strip("'")
    else:
        file_path = input("\n엑셀 파일 경로를 입력하세요: ").strip().strip('"').strip("'")

    if not os.path.exists(file_path):
        print(f"\n오류: 파일을 찾을 수 없습니다 → {file_path}")
        return

    # ── 파일 읽기 ──
    print("\n파일 읽는 중...")
    try:
        df = read_goodauction_xls(file_path)
    except Exception as e:
        print(f"파일 읽기 실패: {e}")
        return

    print(f"총 {len(df)}건 로드")

    # ── 주소 컬럼 탐지 ──
    addr_col = find_address_col(df)
    if not addr_col:
        print("\n[컬럼 목록]")
        for i, c in enumerate(df.columns):
            print(f"  {i:3d}. {c}")
        try:
            idx = int(input("주소 컬럼 번호: ").strip())
            addr_col = df.columns[idx]
        except (ValueError, IndexError):
            print("잘못된 입력입니다.")
            return

    print(f"주소 컬럼: 『{addr_col}』\n")

    # ── 물건별 실거래가 조회 ──
    rows_out = []
    auction_col_count = len(df.columns) + 2  # 법정동코드 + 실거래상태 포함

    for i, row in df.iterrows():
        address = str(row.get(addr_col, '')).strip()
        if not address or address.lower() in ('nan', 'none', ''):
            continue

        seq = i + 1
        print(f"[{seq}/{len(df)}] {address[:70]}")

        lawd_cd, sigungu = get_lawd_cd(address)

        if not lawd_cd:
            print("  → 법정동코드 추출 실패 (시군구를 인식하지 못함)")
            base = row.to_dict()
            base.update({'법정동코드': '-', '실거래_상태': '법정동코드 추출 실패'})
            rows_out.append(base)
            continue

        print(f"  → 법정동코드: {lawd_cd}  ({sigungu})")

        all_trades = fetch_recent_trades(lawd_cd)
        print(f"  → 수신 {len(all_trades)}건", end='')

        # 동 필터링 → 없으면 전 건 사용
        nearby = filter_by_dong(all_trades, address)
        if not nearby:
            nearby = all_trades

        nearby = sort_trades(nearby)[:MAX_TRADES_PER_ITEM]
        print(f"  (인근 {len(nearby)}건 사용)")

        if not nearby:
            base = row.to_dict()
            base.update({'법정동코드': lawd_cd, '실거래_상태': '거래내역 없음'})
            rows_out.append(base)
            continue

        for trade in nearby:
            base = row.to_dict()
            base['법정동코드'] = lawd_cd
            base['실거래_상태'] = '조회성공'
            base.update(fmt_trade(trade))
            rows_out.append(base)

    if not rows_out:
        print("\n저장할 결과가 없습니다.")
        return

    # ── 저장 ──
    result_df = pd.DataFrame(rows_out)
    base_name = os.path.splitext(file_path)[0]
    out_path = f"{base_name}_실거래가_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    print(f"\n저장 중... {out_path}")
    try:
        save_excel(result_df, out_path, auction_col_count)
        print(f"완료!  총 {len(result_df)}행  →  {out_path}")
    except Exception as e:
        print(f"저장 실패: {e}")
        # 서식 없이 재시도
        result_df.to_excel(out_path, index=False)
        print(f"서식 없이 저장됨: {out_path}")


if __name__ == '__main__':
    main()
